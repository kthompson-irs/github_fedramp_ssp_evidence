#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import re
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

DARK = "1F4E78"
LIGHT = "D9EAF7"
GREEN = "E8F5E9"
AMBER = "FFF4D6"
RED = "FDE8E8"
WHITE = "FFFFFF"
thin = Side(style="thin", color="D1D5DB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

INVALID_SHEET_CHARS = r"[\[\]\:\*\?\/\\]"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build POA&M artifacts from blocking findings.")
    parser.add_argument("--input-dir", dest="input_dir", default=None)
    parser.add_argument("--output-dir", dest="output_dir", default=None)
    parser.add_argument("--input", dest="input_legacy", default=None)
    parser.add_argument("--output", dest="output_legacy", default=None)
    args = parser.parse_args()
    args.input_dir = args.input_dir or args.input_legacy or "artifacts/sa-04-10"
    args.output_dir = args.output_dir or args.output_legacy or "poam"
    return args


def read_json(path: Path, default: Any = None) -> Any:
    if not path.exists():
        return default
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return default


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def due_date_for_severity(severity: str) -> str:
    sev = (severity or "").lower()
    if sev in {"high", "critical"}:
        return "30 days"
    if sev in {"moderate", "medium"}:
        return "90 days"
    return "120 days"


def csv_quote(value: str) -> str:
    if any(ch in value for ch in [",", '"', "\n"]):
        return '"' + value.replace('"', '""') + '"'
    return value


def severity_rank(value: str) -> int:
    return {"low": 1, "medium": 2, "moderate": 2, "high": 3, "critical": 4}.get((value or "").lower(), 0)


def normalize_org_repo(finding: Dict[str, Any]) -> Tuple[str, str, str]:
    org = str(finding.get("organization") or "")
    repo = str(finding.get("repository") or "")
    repo_full = str(finding.get("repository_full") or "")

    if not repo_full and org and repo:
        repo_full = f"{org}/{repo}"

    if not org and repo_full and "/" in repo_full:
        org = repo_full.split("/", 1)[0]

    if not repo and repo_full and "/" in repo_full:
        repo = repo_full.split("/", 1)[1]

    return org, repo, repo_full


def load_enterprise_orgs(input_dir: Path, findings: List[Dict[str, Any]]) -> List[str]:
    """
    Prefer live inventory copied into the package build outputs; fall back to unique orgs in findings.
    """
    inv_path = input_dir / "enterprise_organizations.json"
    if inv_path.exists():
        inv = read_json(inv_path, {})
        if isinstance(inv, dict):
            orgs = inv.get("organizations", []) or []
            names = [str(o.get("slug") or o.get("display_name") or o.get("name") or "").strip() for o in orgs]
            names = [n for n in names if n]
            if names:
                return sorted(dict.fromkeys(names).keys())

    orgs = []
    for finding in findings:
        org, _, _ = normalize_org_repo(finding)
        if org:
            orgs.append(org)
    return sorted(dict.fromkeys(orgs).keys())


def build_rows(findings: List[Dict[str, Any]], default_org: str = "", default_repo: str = "") -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    for idx, finding in enumerate(findings, start=1):
        severity = str(finding.get("severity", "")).strip()
        organization, repository, repository_full = normalize_org_repo(finding)
        organization = organization or default_org
        repository = repository or default_repo
        if not repository_full and organization and repository:
            repository_full = f"{organization}/{repository}"

        rows.append(
            {
                "poam_id": f"POAM-{idx:03d}",
                "category": str(finding.get("category", "")),
                "organization": organization,
                "repository": repository,
                "repository_full": repository_full,
                "identifier": str(finding.get("identifier", "")),
                "title": str(finding.get("title", "")),
                "severity": severity,
                "source_url": str(finding.get("html_url", "")),
                "status": "open",
                "recommended_due_date": due_date_for_severity(severity),
                "remediation_owner": "TBD",
                "tracking_reference": "",
                "notes": "Auto-generated from blocking SA-04(10) findings",
                "generated_at": utc_now(),
            }
        )
    return rows


def write_csv(path: Path, rows: List[Dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "poam_id",
        "category",
        "organization",
        "repository",
        "repository_full",
        "identifier",
        "title",
        "severity",
        "source_url",
        "status",
        "recommended_due_date",
        "remediation_owner",
        "tracking_reference",
        "notes",
        "generated_at",
    ]
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def safe_sheet_name(name: str, used: set[str]) -> str:
    cleaned = re.sub(INVALID_SHEET_CHARS, "-", name).strip()
    cleaned = cleaned[:31] or "Sheet"
    candidate = cleaned
    idx = 2
    while candidate in used:
        suffix = f"-{idx}"
        candidate = cleaned[: 31 - len(suffix)] + suffix
        idx += 1
    used.add(candidate)
    return candidate


def style_header(ws, header_row: int, headers: List[str]) -> None:
    for idx, header in enumerate(headers, 1):
        cell = ws.cell(header_row, idx, header)
        cell.font = Font(color=WHITE, bold=True)
        cell.fill = PatternFill("solid", fgColor=DARK)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def create_table(ws, start_row: int, end_row: int, end_col: int, table_name: str) -> None:
    if end_row < start_row:
        return
    ref = f"A{start_row}:{chr(64 + end_col)}{end_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def write_org_sheet(ws, org_name: str, rows: List[Dict[str, str]]) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    headers = [
        "POAM ID",
        "Category",
        "Organization",
        "Repository",
        "Repository Full",
        "Identifier",
        "Title",
        "Severity",
        "Source URL",
        "Status",
        "Due Date",
        "Owner",
        "Tracking Ref",
        "Notes",
        "Generated At",
    ]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"] = f"POA&M — {org_name}"
    ws["A1"].font = Font(color=WHITE, bold=True, size=13)
    ws["A1"].fill = PatternFill("solid", fgColor=DARK)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws["A2"] = f"Only findings for {org_name} are shown on this tab."
    ws["A2"].fill = PatternFill("solid", fgColor=LIGHT)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    style_header(ws, 3, headers)

    for r_idx, row in enumerate(rows, start=4):
        values = [
            row["poam_id"],
            row["category"],
            row["organization"],
            row["repository"],
            row["repository_full"],
            row["identifier"],
            row["title"],
            row["severity"],
            row["source_url"],
            row["status"],
            row["recommended_due_date"],
            row["remediation_owner"],
            row["tracking_reference"],
            row["notes"],
            row["generated_at"],
        ]
        for c_idx, value in enumerate(values, start=1):
            cell = ws.cell(r_idx, c_idx, value)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="left" if c_idx in {2, 3, 4, 5, 6, 7, 9, 12, 13, 14, 15} else "center",
                vertical="center",
                wrap_text=True,
            )

    widths = [14, 16, 18, 24, 30, 24, 34, 12, 44, 12, 12, 18, 18, 30, 24]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width

    if rows:
        create_table(ws, 3, len(rows) + 3, len(headers), f"poam_{re.sub(r'[^A-Za-z0-9]', '_', org_name)[:20].lower()}")


def write_overview_sheet(ws, rows: List[Dict[str, str]], orgs: List[str]) -> None:
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:D1")
    ws["A1"] = "SA-04(10) POA&M Overview"
    ws["A1"].font = Font(color=WHITE, bold=True, size=13)
    ws["A1"].fill = PatternFill("solid", fgColor=DARK)
    ws["A1"].alignment = Alignment(horizontal="center")

    total = len(rows)
    by_org = defaultdict(int)
    for row in rows:
        by_org[row["organization"]] += 1

    summary_rows = [
        ["Generated At", utc_now()],
        ["Total Findings", total],
        ["Organizations in Inventory", len(orgs)],
        ["Organizations with Findings", len(by_org)],
    ]
    for i, (k, v) in enumerate(summary_rows, start=3):
        ws.cell(i, 1, k).font = Font(bold=True)
        ws.cell(i, 1).fill = PatternFill("solid", fgColor=LIGHT)
        ws.cell(i, 1).border = border
        ws.cell(i, 2, v).border = border

    start = 9
    ws.cell(start, 1, "Organization").font = Font(color=WHITE, bold=True)
    ws.cell(start, 1).fill = PatternFill("solid", fgColor=DARK)
    ws.cell(start, 2, "Findings").font = Font(color=WHITE, bold=True)
    ws.cell(start, 2).fill = PatternFill("solid", fgColor=DARK)
    ws.cell(start, 3, "Coverage").font = Font(color=WHITE, bold=True)
    ws.cell(start, 3).fill = PatternFill("solid", fgColor=DARK)

    for idx, org in enumerate(orgs, start=start + 1):
        count = by_org.get(org, 0)
        ws.cell(idx, 1, org).border = border
        ws.cell(idx, 2, count).border = border
        ws.cell(idx, 3, "Yes" if count else "No").border = border

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 24


def write_diagnostics_sheet(ws, diagnostics: Dict[str, Any]) -> None:
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:C1")
    ws["A1"] = "Collector Diagnostics"
    ws["A1"].font = Font(color=WHITE, bold=True, size=13)
    ws["A1"].fill = PatternFill("solid", fgColor=DARK)
    ws["A1"].alignment = Alignment(horizontal="center")

    overview = [
        ["Generated At", diagnostics.get("generated_at", "")],
        ["Scope", diagnostics.get("scope", "")],
        ["Enterprise", diagnostics.get("enterprise", "")],
        ["Organization", diagnostics.get("organization", "")],
        ["Repository", diagnostics.get("repository", "")],
        ["Repository Full", diagnostics.get("repository_full", "")],
        ["Selected Auth Kind", diagnostics.get("selected_auth_kind", "")],
        ["Notes", "; ".join(diagnostics.get("notes", []) or [])],
    ]
    for i, (k, v) in enumerate(overview, 3):
        ws.cell(i, 1, k).font = Font(bold=True)
        ws.cell(i, 1).fill = PatternFill("solid", fgColor=LIGHT)
        ws.cell(i, 1).border = border
        ws.cell(i, 2, v).border = border
        ws.cell(i, 2).alignment = Alignment(wrap_text=True)

    token_row = 13
    ws.merge_cells(start_row=token_row, start_column=1, end_row=token_row, end_column=8)
    ws.cell(token_row, 1, "Token Attempts").font = Font(color=WHITE, bold=True)
    ws.cell(token_row, 1).fill = PatternFill("solid", fgColor=DARK)
    ws.cell(token_row, 1).alignment = Alignment(horizontal="center")

    headers = ["Token Env", "Auth Kind", "Status", "Login", "Type", "Accepted Permissions", "SSO", "Message"]
    header_row = token_row + 1
    style_header(ws, header_row, headers)

    r = header_row + 1
    for attempt in diagnostics.get("token_attempts", []) or []:
        identity = attempt.get("identity") or {}
        row = [
            attempt.get("token_env", ""),
            attempt.get("auth_kind", ""),
            attempt.get("status", ""),
            identity.get("login", ""),
            identity.get("type", ""),
            identity.get("accepted_permissions", ""),
            identity.get("sso", ""),
            identity.get("message", "") or "",
        ]
        for c_idx, val in enumerate(row, 1):
            c = ws.cell(r, c_idx, val)
            c.border = border
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        r += 1

    widths = [24, 26, 18, 18, 14, 24, 20, 80]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width


def main() -> int:
    args = parse_args()
    input_dir = Path(args.input_dir)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    findings = read_json(input_dir / "blocking_findings.json", []) or []
    if not isinstance(findings, list):
        raise SystemExit("blocking_findings.json is malformed or not a list")

    summary = read_json(input_dir / "summary.json", {}) or {}
    diagnostics = read_json(input_dir / "diagnostics.json", {}) or {}
    default_org = str(summary.get("organization") or summary.get("owner") or "")
    repo_full = str(summary.get("repository") or "")
    default_repo = repo_full.split("/", 1)[1] if "/" in repo_full else repo_full

    rows = build_rows(findings, default_org=default_org, default_repo=default_repo)
    generated_at = utc_now()

    orgs = load_enterprise_orgs(input_dir, findings)
    orgs = orgs or sorted({row["organization"] for row in rows if row["organization"]})
    if summary.get("scope") == "enterprise" and summary.get("enterprise"):
        # keep inventory sheet in enterprise workbooks even when there are no findings
        pass

    write_csv(output_dir / "poam.csv", rows)

    wb = Workbook()
    overview = wb.active
    overview.title = "Overview"
    write_overview_sheet(overview, rows, orgs)

    diag = wb.create_sheet("Diagnostics")
    write_diagnostics_sheet(diag, diagnostics or {"generated_at": generated_at})

    used_sheet_names: set[str] = {"Overview", "Diagnostics"}

    # Master sheet
    master = wb.create_sheet("POA&M")
    master.sheet_view.showGridLines = False
    master.freeze_panes = "A4"
    master_headers = [
        "POAM ID",
        "Category",
        "Organization",
        "Repository",
        "Repository Full",
        "Identifier",
        "Title",
        "Severity",
        "Source URL",
        "Status",
        "Due Date",
        "Owner",
        "Tracking Ref",
        "Notes",
        "Generated At",
    ]
    master.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(master_headers))
    master["A1"] = "SA-04(10) Auto-generated POA&M"
    master["A1"].font = Font(color=WHITE, bold=True, size=14)
    master["A1"].fill = PatternFill("solid", fgColor=DARK)
    master["A1"].alignment = Alignment(horizontal="center")
    master.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(master_headers))
    master["A2"] = f"Generated {generated_at}"
    master["A2"].fill = PatternFill("solid", fgColor=LIGHT)
    master["A2"].alignment = Alignment(horizontal="left")
    style_header(master, 3, master_headers)

    for r_idx, row in enumerate(rows, start=4):
        values = [
            row["poam_id"],
            row["category"],
            row["organization"],
            row["repository"],
            row["repository_full"],
            row["identifier"],
            row["title"],
            row["severity"],
            row["source_url"],
            row["status"],
            row["recommended_due_date"],
            row["remediation_owner"],
            row["tracking_reference"],
            row["notes"],
            row["generated_at"],
        ]
        for c_idx, value in enumerate(values, start=1):
            cell = master.cell(r_idx, c_idx, value)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="left" if c_idx in {2, 3, 4, 5, 6, 7, 9, 12, 13, 14} else "center",
                vertical="center",
                wrap_text=True,
            )
            if c_idx == 8:
                sev = str(value).lower()
                if sev in {"critical", "high"}:
                    cell.fill = PatternFill("solid", fgColor=RED)
                elif sev in {"medium", "moderate"}:
                    cell.fill = PatternFill("solid", fgColor=AMBER)
                elif sev == "low":
                    cell.fill = PatternFill("solid", fgColor=GREEN)

    widths = [14, 16, 18, 24, 30, 24, 34, 12, 44, 12, 12, 18, 18, 30, 24]
    for idx, width in enumerate(widths, start=1):
        master.column_dimensions[chr(64 + idx)].width = width
    create_table(master, 3, len(rows) + 3, len(master_headers), "poam_master")

    # One tab per org, with only that org's findings
    by_org: Dict[str, List[Dict[str, str]]] = defaultdict(list)
    for row in rows:
        by_org[row["organization"]].append(row)

    for org in orgs:
        org_rows = sorted(by_org.get(org, []), key=lambda r: (-severity_rank(r["severity"]), r["repository_full"], r["identifier"]))
        tab = safe_sheet_name(org, used_sheet_names)
        ws = wb.create_sheet(tab)
        write_org_sheet(ws, org, org_rows)
        if not org_rows:
            ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=15)
            ws["A5"] = "No findings for this organization in the current 30-day window."
            ws["A5"].alignment = Alignment(horizontal="center")
            ws["A5"].font = Font(italic=True)

    # Any orgs found only in findings and not inventory
    for org in sorted(set(by_org.keys()) - set(orgs)):
        org_rows = sorted(by_org.get(org, []), key=lambda r: (-severity_rank(r["severity"]), r["repository_full"], r["identifier"]))
        tab = safe_sheet_name(org, used_sheet_names)
        ws = wb.create_sheet(tab)
        write_org_sheet(ws, org, org_rows)

    wb.save(output_dir / "poam.xlsx")

    # Markdown summary for quick review
    md_lines = [
        "# SA-04(10) POA&M",
        "",
        f"- Generated at: {generated_at}",
        f"- Finding count: {len(rows)}",
        f"- Organization count: {len(orgs)}",
        "",
        "| POAM ID | Organization | Repository | Category | Identifier | Severity | Due Date | Status |",
        "|---|---|---|---|---|---|---|---|",
    ]
    for row in rows:
        md_lines.append(
            f"| {row['poam_id']} | {row['organization']} | {row['repository']} | {row['category']} | "
            f"{row['identifier']} | {row['severity']} | {row['recommended_due_date']} | {row['status']} |"
        )
    (output_dir / "poam.md").write_text("\n".join(md_lines) + "\n", encoding="utf-8")

    print(f"POA&M generated with {len(rows)} row(s) in {output_dir}")
    print(f"Organization tabs: {len(orgs)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
