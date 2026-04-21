#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


DARK = "1F4E78"
LIGHT = "D9EAF7"
GREEN = "E8F5E9"
AMBER = "FFF4D6"
RED = "FDE8E8"
WHITE = "FFFFFF"
thin = Side(style="thin", color="D1D5DB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build enterprise SA-04(10) 30-day spreadsheets.")

    parser.add_argument("--input-dir", dest="input_dir", default=None)
    parser.add_argument("--output-dir", dest="output_dir", default=None)

    parser.add_argument("--input", dest="input_legacy", default=None)
    parser.add_argument("--output", dest="output_legacy", default=None)

    args = parser.parse_args()
    args.input_dir = args.input_dir or args.input_legacy or "artifacts/sa-04-10"
    args.output_dir = args.output_dir or args.output_legacy or "spreadsheets"
    return args


def read_json(path: Path, default: Any = None) -> Any:
    if not path.exists():
        return default
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return default


def safe_dt(value: Optional[str]) -> Optional[datetime]:
    if not value:
        return None
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00"))
    except Exception:
        return None


def utc_date_from_iso(value: Optional[str]) -> Optional[str]:
    dt = safe_dt(value)
    return dt.date().isoformat() if dt else None


def repo_parts(repository_full: str) -> tuple[str, str]:
    if "/" in repository_full:
        org, repo = repository_full.split("/", 1)
        return org, repo
    return "", repository_full


def read_history(input_dir: Path) -> List[Dict[str, Any]]:
    history_dir = input_dir / "history"
    records: List[Dict[str, Any]] = []

    history_jsonl = history_dir / "history.jsonl"
    if history_jsonl.exists():
        for line in history_jsonl.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line:
                continue
            try:
                record = json.loads(line)
                if isinstance(record, dict):
                    records.append(record)
            except json.JSONDecodeError:
                continue

    if not records:
        current = read_json(input_dir / "current_snapshot.json")
        if isinstance(current, dict) and current:
            records.append(current)
        else:
            summary = read_json(input_dir / "summary.json")
            if isinstance(summary, dict) and summary:
                records.append(summary)

    per_day: Dict[str, Dict[str, Any]] = {}
    for rec in records:
        day = rec.get("date") or utc_date_from_iso(rec.get("generated_at")) or ""
        if day:
            per_day[day] = rec

    return [per_day[k] for k in sorted(per_day.keys())]


def last_30_days(history: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return history[-30:] if len(history) > 30 else history


def severity_rank(value: str) -> int:
    return {"low": 1, "medium": 2, "moderate": 2, "high": 3, "critical": 4}.get((value or "").lower(), 0)


def normalize_severity(value: Optional[str]) -> str:
    if not value:
        return "unknown"
    return str(value).strip().lower()


def snapshot_meta(snapshot: Dict[str, Any]) -> Dict[str, str]:
    repository_full = str(snapshot.get("repository_full") or snapshot.get("repository") or "")
    organization = str(snapshot.get("organization") or snapshot.get("owner") or "")
    repository = str(snapshot.get("repository") or "")

    if not organization or not repository:
        if "/" in repository_full:
            org, repo = repo_parts(repository_full)
            organization = organization or org
            repository = repository or repo

    if not repository_full and organization and repository:
        repository_full = f"{organization}/{repository}"

    return {
        "organization": organization,
        "repository": repository,
        "repository_full": repository_full,
    }


def findings_from_snapshot(snapshot: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    Normalize all alert entries from one snapshot into spreadsheet-ready rows.
    """
    entries: List[Dict[str, Any]] = []
    results = snapshot.get("results", {}) or {}
    date = snapshot.get("date") or utc_date_from_iso(snapshot.get("generated_at")) or ""
    scope = snapshot.get("scope") or ""
    meta = snapshot_meta(snapshot)
    organization = meta["organization"]
    repository = meta["repository"]
    repository_full = meta["repository_full"]

    # Code scanning alerts
    for alert in (results.get("code_scanning", {}) or {}).get("alerts", []) or []:
        rule = alert.get("rule") or {}
        severity = normalize_severity(rule.get("severity"))
        loc = (alert.get("most_recent_instance") or {}).get("location") or {}
        alert_meta = snapshot_meta(alert)
        entries.append(
            {
                "date": date,
                "scope": scope,
                "organization": alert_meta["organization"] or organization,
                "repository": alert_meta["repository"] or repository,
                "repository_full": alert_meta["repository_full"] or repository_full,
                "category": "CodeQL",
                "severity": severity,
                "severity_rank": severity_rank(severity),
                "status": str(alert.get("state") or "open"),
                "identifier": rule.get("id") or "unknown-rule",
                "title": rule.get("name") or "unknown-title",
                "source": "code_scanning",
                "source_url": alert.get("html_url") or "",
                "path": loc.get("path") or "",
                "manifest": "",
                "evidence_ref": snapshot.get("generated_at") or "",
                "notes": "Historical CodeQL alert",
            }
        )

    # Dependabot alerts
    for alert in (results.get("dependabot", {}) or {}).get("alerts", []) or []:
        advisory = alert.get("security_advisory") or {}
        severity = normalize_severity(advisory.get("severity"))
        dependency = (((alert.get("dependency") or {}).get("package") or {}).get("name")) or "unknown-package"
        alert_meta = snapshot_meta(alert)
        entries.append(
            {
                "date": date,
                "scope": scope,
                "organization": alert_meta["organization"] or organization,
                "repository": alert_meta["repository"] or repository,
                "repository_full": alert_meta["repository_full"] or repository_full,
                "category": "Dependabot",
                "severity": severity,
                "severity_rank": severity_rank(severity),
                "status": str(alert.get("state") or "open"),
                "identifier": advisory.get("ghsa_id") or advisory.get("cve_id") or "unknown-advisory",
                "title": advisory.get("summary") or "unknown-advisory",
                "source": "dependabot",
                "source_url": alert.get("html_url") or "",
                "path": dependency,
                "manifest": (((alert.get("dependency") or {}).get("manifest_path")) or ""),
                "evidence_ref": snapshot.get("generated_at") or "",
                "notes": "Historical Dependabot alert",
            }
        )

    # Secret scanning alerts
    for alert in (results.get("secret_scanning", {}) or {}).get("alerts", []) or []:
        severity = "high" if str(alert.get("state") or "").lower() == "open" else "low"
        alert_meta = snapshot_meta(alert)
        entries.append(
            {
                "date": date,
                "scope": scope,
                "organization": alert_meta["organization"] or organization,
                "repository": alert_meta["repository"] or repository,
                "repository_full": alert_meta["repository_full"] or repository_full,
                "category": "Secret Scanning",
                "severity": severity,
                "severity_rank": severity_rank(severity),
                "status": str(alert.get("state") or "open"),
                "identifier": alert.get("secret_type") or alert.get("secret_type_display_name") or "unknown-secret",
                "title": alert.get("secret_type_display_name") or alert.get("secret_type") or "unknown-secret",
                "source": "secret_scanning",
                "source_url": alert.get("html_url") or "",
                "path": "",
                "manifest": "",
                "evidence_ref": snapshot.get("generated_at") or "",
                "notes": "Historical secret scanning alert",
            }
        )

    return entries


def sort_entries(entries: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Highest severity first, then newest date, then organization/repository, then identifier.
    """
    return sorted(
        entries,
        key=lambda item: (
            -int(item.get("severity_rank", 0)),
            str(item.get("date", "")),
            str(item.get("organization", "")),
            str(item.get("repository", "")),
            str(item.get("category", "")),
            str(item.get("identifier", "")),
        ),
    )


def write_table_sheet(
    ws,
    title: str,
    subtitle: str,
    headers: List[str],
    rows: List[List[Any]],
    table_name: str,
    widths: List[int],
) -> None:
    cols = len(headers)
    end_col = chr(64 + cols)

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    ws["A1"] = title
    ws["A1"].font = Font(color=WHITE, bold=True, size=14)
    ws["A1"].fill = PatternFill("solid", fgColor=DARK)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
    ws["A2"] = subtitle
    ws["A2"].fill = PatternFill("solid", fgColor=LIGHT)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for idx, header in enumerate(headers, 1):
        c = ws.cell(3, idx, header)
        c.font = Font(color=WHITE, bold=True)
        c.fill = PatternFill("solid", fgColor=DARK)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    for r_idx, row in enumerate(rows, 4):
        for c_idx, value in enumerate(row, 1):
            c = ws.cell(r_idx, c_idx, value)
            c.border = border
            c.alignment = Alignment(
                horizontal="left" if c_idx in {3, 4, 5, 8, 9, 10, 11, 12, 13, 14} else "center",
                vertical="center",
                wrap_text=True,
            )
            if c_idx == 1 and value:
                c.number_format = "yyyy-mm-dd"

            if c_idx == 7:  # severity
                sev = str(value).lower()
                if sev in {"critical", "high"}:
                    c.fill = PatternFill("solid", fgColor=RED)
                elif sev in {"medium", "moderate"}:
                    c.fill = PatternFill("solid", fgColor=AMBER)
                elif sev == "low":
                    c.fill = PatternFill("solid", fgColor=GREEN)

    if rows:
        ref = f"A3:{end_col}{len(rows) + 3}"
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = width


def write_summary_sheet(ws, title: str, rows: List[List[Any]]) -> None:
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:B1")
    ws["A1"] = title
    ws["A1"].font = Font(color=WHITE, bold=True, size=13)
    ws["A1"].fill = PatternFill("solid", fgColor=DARK)
    ws["A1"].alignment = Alignment(horizontal="center")

    for i, (k, v) in enumerate(rows, 3):
        ws.cell(i, 1, k).font = Font(bold=True)
        ws.cell(i, 1).fill = PatternFill("solid", fgColor=LIGHT)
        ws.cell(i, 1).border = border
        ws.cell(i, 2, v).border = border
        ws.cell(i, 2).alignment = Alignment(wrap_text=True)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 95


def build_workbook(
    path: Path,
    sheet_title: str,
    rows: List[List[Any]],
    summary_rows: List[List[Any]],
    notes: List[str],
    table_name: str,
    widths: List[int],
    headers: List[str],
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "30-Day Log"

    write_table_sheet(
        ws=ws,
        title=sheet_title,
        subtitle="Enterprise-wide historical alert entries aggregated from artifacts/sa-04-10/history. No placeholder rows are inserted.",
        headers=headers,
        rows=rows,
        table_name=table_name,
        widths=widths,
    )

    summary = wb.create_sheet("Snapshot Summary")
    write_summary_sheet(summary, f"{sheet_title} Summary", summary_rows)

    readme = wb.create_sheet("Read Me")
    readme.sheet_view.showGridLines = False
    readme.merge_cells("A1:B1")
    readme["A1"] = f"{sheet_title} — Audit Notes"
    readme["A1"].font = Font(color=WHITE, bold=True, size=13)
    readme["A1"].fill = PatternFill("solid", fgColor=DARK)

    for i, note in enumerate(notes, 3):
        readme.cell(i, 1, f"Note {i-2}").font = Font(bold=True)
        readme.cell(i, 1).fill = PatternFill("solid", fgColor=LIGHT)
        readme.cell(i, 1).border = border
        readme.cell(i, 2, note).border = border
        readme.cell(i, 2).alignment = Alignment(wrap_text=True)

    readme.column_dimensions["A"].width = 18
    readme.column_dimensions["B"].width = 95

    wb.save(path)


def aggregate_entries(history: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    entries: List[Dict[str, Any]] = []
    for snapshot in history:
        entries.extend(findings_from_snapshot(snapshot))
    return entries


def make_sheet_rows(entries: List[Dict[str, Any]]) -> Dict[str, List[List[Any]]]:
    dep_rows: List[List[Any]] = []
    code_rows: List[List[Any]] = []
    sec_rows: List[List[Any]] = []

    for e in entries:
        row = [
            e.get("date", ""),
            e.get("scope", ""),
            e.get("organization", ""),
            e.get("repository", ""),
            e.get("repository_full", ""),
            e.get("category", ""),
            e.get("severity", ""),
            e.get("status", ""),
            e.get("identifier", ""),
            e.get("title", ""),
            e.get("source_url", ""),
            e.get("path", ""),
            e.get("manifest", ""),
            e.get("evidence_ref", ""),
            e.get("notes", ""),
        ]
        if e.get("source") == "dependabot":
            dep_rows.append(row)
        elif e.get("source") == "code_scanning":
            code_rows.append(row)
        elif e.get("source") == "secret_scanning":
            sec_rows.append(row)

    dep_rows = sorted(dep_rows, key=lambda r: (-severity_rank(str(r[6])), str(r[0]), str(r[2]), str(r[3]), str(r[8])))
    code_rows = sorted(code_rows, key=lambda r: (-severity_rank(str(r[6])), str(r[0]), str(r[2]), str(r[3]), str(r[8])))
    sec_rows = sorted(sec_rows, key=lambda r: (-severity_rank(str(r[6])), str(r[0]), str(r[2]), str(r[3]), str(r[8])))

    security_rows = sorted(
        dep_rows + code_rows + sec_rows,
        key=lambda r: (-severity_rank(str(r[6])), str(r[0]), str(r[2]), str(r[3]), str(r[5]), str(r[8])),
    )

    return {
        "dependabot": dep_rows,
        "codeql": code_rows,
        "security": security_rows,
    }


def enterprise_summary_text(snapshot: Dict[str, Any]) -> str:
    repositories = snapshot.get("repositories", []) or []
    lines = [
        f"Enterprise slug: {snapshot.get('enterprise', '')}",
        f"Repository count: {snapshot.get('repository_count', 0)}",
    ]
    if repositories:
        lines.append("Covered repositories:")
        for repo in repositories:
            lines.append(f"- {repo}")
    return "\n".join(lines)


def main() -> int:
    args = parse_args()
    input_dir = Path(args.input_dir)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    history = read_history(input_dir)
    if not history:
        raise SystemExit("No historical snapshots found in artifacts/sa-04-10")

    entries = sort_entries(aggregate_entries(last_30_days(history)))
    rows = make_sheet_rows(entries)

    latest = history[-1]
    summary_blocking = latest.get("overall", {}).get("blocking_count", 0)

    headers = [
        "Date",
        "Scope",
        "Organization",
        "Repository",
        "Repository Full",
        "Category",
        "Severity",
        "Status",
        "Identifier",
        "Title",
        "Source URL",
        "Path",
        "Manifest",
        "Evidence Ref",
        "Notes",
    ]
    widths = [12, 16, 18, 24, 30, 16, 12, 12, 24, 34, 44, 30, 24, 24, 30]

    build_workbook(
        path=output_dir / "dependabot_30_day_log.xlsx",
        sheet_title="Dependabot 30-Day Log",
        rows=rows["dependabot"],
        summary_rows=[
            ["Generated At", latest.get("generated_at", "")],
            ["Scope", latest.get("scope", "")],
            ["Organization", latest.get("organization", "")],
            ["Repository", latest.get("repository", "")],
            ["Repository Full", latest.get("repository_full", "")],
            ["Entry Count", len(rows["dependabot"])],
            ["Blocking Findings", summary_blocking],
            ["Status", latest.get("overall", {}).get("status", "").title()],
        ],
        notes=[
            "This workbook lists real Dependabot alert entries from the last 30 days.",
            "Rows are sorted by severity first, then by date and repository.",
            "No synthetic placeholder rows are inserted.",
        ],
        table_name="dependabot_30_day_log",
        widths=widths,
        headers=headers,
    )

    build_workbook(
        path=output_dir / "codeql_30_day_log.xlsx",
        sheet_title="CodeQL 30-Day Log",
        rows=rows["codeql"],
        summary_rows=[
            ["Generated At", latest.get("generated_at", "")],
            ["Scope", latest.get("scope", "")],
            ["Organization", latest.get("organization", "")],
            ["Repository", latest.get("repository", "")],
            ["Repository Full", latest.get("repository_full", "")],
            ["Entry Count", len(rows["codeql"])],
            ["Blocking Findings", summary_blocking],
            ["Status", latest.get("overall", {}).get("status", "").title()],
        ],
        notes=[
            "This workbook lists real CodeQL alert entries from the last 30 days.",
            "Rows are sorted by severity first, then by date and repository.",
            "Evidence references point back to the originating snapshot.",
        ],
        table_name="codeql_30_day_log",
        widths=widths,
        headers=headers,
    )

    build_workbook(
        path=output_dir / "security_30_day_log.xlsx",
        sheet_title="Security 30-Day Log",
        rows=rows["security"],
        summary_rows=[
            ["Generated At", latest.get("generated_at", "")],
            ["Scope", latest.get("scope", "")],
            ["Organization", latest.get("organization", "")],
            ["Repository", latest.get("repository", "")],
            ["Repository Full", latest.get("repository_full", "")],
            ["Entry Count", len(rows["security"])],
            ["Blocking Findings", summary_blocking],
            ["Collection Errors", latest.get("overall", {}).get("error_count", 0)],
            ["Status", latest.get("overall", {}).get("status", "").title()],
        ],
        notes=[
            "This workbook combines CodeQL, Dependabot, and Secret Scanning entries from the last 30 days.",
            "Entries are sorted by severity first, then by date and repository.",
            "Use this workbook as the aggregate enterprise security evidence log.",
        ],
        table_name="security_30_day_log",
        widths=widths,
        headers=headers,
    )

    if latest.get("scope") == "enterprise":
        write_path = output_dir / "enterprise_coverage.txt"
        write_path.write_text(enterprise_summary_text(latest) + "\n", encoding="utf-8")

    print(f"Spreadsheets generated in {output_dir}")
    print(f"Dependabot rows: {len(rows['dependabot'])}")
    print(f"CodeQL rows: {len(rows['codeql'])}")
    print(f"Security rows: {len(rows['security'])}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
