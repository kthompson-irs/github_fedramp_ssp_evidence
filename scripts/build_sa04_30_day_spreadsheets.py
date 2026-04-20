#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import shutil
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


DARK = "1F4E78"
LIGHT = "D9EAF7"
GREEN = "E8F5E9"
AMBER = "FFF4D6"
RED = "FDE8E8"
WHITE = "FFFFFF"
thin = Side(style="thin", color="D1D5DB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build SA-04(10) 30-day spreadsheets.")

    parser.add_argument("--input-dir", dest="input_dir", default=None)
    parser.add_argument("--output-dir", dest="output_dir", default=None)

    parser.add_argument("--input", dest="input_legacy", default=None)
    parser.add_argument("--output", dest="output_legacy", default=None)

    args = parser.parse_args()
    args.input_dir = args.input_dir or args.input_legacy or "artifacts/sa-04-10"
    args.output_dir = args.output_dir or args.output_legacy or "spreadsheets"
    return args


def read_json(path: Path, default: Any = None) -> Any:
    if not path.exists():
        return default
    return json.loads(path.read_text(encoding="utf-8"))


def utc_date_from_iso(value: str) -> Optional[datetime.date]:
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00")).date()
    except Exception:
        return None


def latest_snapshot(input_dir: Path) -> Dict[str, Any]:
    current = read_json(input_dir / "current_snapshot.json")
    if isinstance(current, dict) and current:
        return current
    summary = read_json(input_dir / "summary.json")
    if isinstance(summary, dict) and summary:
        return summary
    raise SystemExit("summary.json or current_snapshot.json not found or empty")


def read_history(input_dir: Path) -> List[Dict[str, Any]]:
    history_dir = input_dir / "history"
    items: List[Dict[str, Any]] = []

    jsonl = history_dir / "history.jsonl"
    if jsonl.exists():
        for line in jsonl.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line:
                try:
                    items.append(json.loads(line))
                except json.JSONDecodeError:
                    continue

    if not items:
        items.append(latest_snapshot(input_dir))

    # keep most recent snapshot per day
    dedup: Dict[str, Dict[str, Any]] = {}
    for item in items:
        day = item.get("date") or (item.get("generated_at") or "")[:10]
        if day:
            dedup[day] = item

    return [dedup[k] for k in sorted(dedup.keys())]


def last_30_days(history: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return history[-30:] if len(history) > 30 else history


def tableize(ws, ref: str, name: str) -> None:
    t = Table(displayName=name, ref=ref)
    t.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(t)


def style_header(ws, row: int, headers: List[str]) -> None:
    for idx, h in enumerate(headers, 1):
        c = ws.cell(row, idx, h)
        c.font = Font(color=WHITE, bold=True)
        c.fill = PatternFill("solid", fgColor=DARK)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border


def apply_common_sheet_layout(ws, title: str, subtitle: str, header_row: int, table_name: str, rows: List[List[Any]], widths: List[int]) -> None:
    cols = len(widths)
    end_col = chr(64 + cols)

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = f"A{header_row + 1}"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    ws["A1"] = title
    ws["A1"].font = Font(color=WHITE, bold=True, size=14)
    ws["A1"].fill = PatternFill("solid", fgColor=DARK)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
    ws["A2"] = subtitle
    ws["A2"].fill = PatternFill("solid", fgColor=LIGHT)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    if rows:
        table_ref = f"A{header_row}:{end_col}{header_row + len(rows)}"
        tableize(ws, table_ref, table_name)

    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = width


def snapshot_summary_rows(kind: str, snapshot: Dict[str, Any]) -> List[List[Any]]:
    results = snapshot.get("results", {})
    overall = snapshot.get("overall", {})

    if kind == "dependabot":
        dep = results.get("dependabot", {})
        return [
            ["Generated At", snapshot.get("generated_at", "")],
            ["Date", snapshot.get("date", "")],
            ["Scope", snapshot.get("scope", "")],
            ["Repository", snapshot.get("repository", "")],
            ["Open Alerts", dep.get("count", 0)],
            ["Blocking Findings", dep.get("blocking_count", 0)],
            ["Accessible", "Yes" if dep.get("accessible") else "No"],
            ["Skipped", "Yes" if dep.get("skipped") else "No"],
            ["Skip Reason", dep.get("skip_reason") or ""],
            ["Status", "Open" if dep.get("blocking_count", 0) else "Clear"],
        ]

    if kind == "codeql":
        code = results.get("code_scanning", {})
        return [
            ["Generated At", snapshot.get("generated_at", "")],
            ["Date", snapshot.get("date", "")],
            ["Scope", snapshot.get("scope", "")],
            ["Repository", snapshot.get("repository", "")],
            ["Open Alerts", code.get("count", 0)],
            ["Blocking Findings", code.get("blocking_count", 0)],
            ["Accessible", "Yes" if code.get("accessible") else "No"],
            ["Skipped", "Yes" if code.get("skipped") else "No"],
            ["Skip Reason", code.get("skip_reason") or ""],
            ["Status", "Open" if code.get("blocking_count", 0) else "Clear"],
        ]

    sec = results.get("secret_scanning", {})
    return [
        ["Generated At", snapshot.get("generated_at", "")],
        ["Date", snapshot.get("date", "")],
        ["Scope", snapshot.get("scope", "")],
        ["Repository", snapshot.get("repository", "")],
        ["Open Code Scanning", results.get("code_scanning", {}).get("count", 0)],
        ["Open Dependabot", results.get("dependabot", {}).get("count", 0)],
        ["Open Secret Scanning", sec.get("count", 0)],
        ["Blocking Findings", overall.get("blocking_count", 0)],
        ["Collection Errors", overall.get("error_count", 0)],
        ["Status", overall.get("status", "").title()],
    ]


def build_workbook(path: Path, title: str, kind: str, rows: List[List[Any]], snapshot: Dict[str, Any]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "30-Day Log"

    if kind == "security":
        headers = ["Date", "Scope", "Repository", "Code Scanning", "Dependabot", "Secret Scanning", "Blocking Findings", "Status", "Evidence Ref", "Notes"]
        widths = [12, 16, 44, 14, 14, 16, 16, 14, 26, 34]
        row_len = 10
    else:
        headers = ["Date", "Scope", "Repository", "Open Alerts", "Blocking Findings", "Highest Severity", "Status", "Evidence Ref", "Notes"]
        widths = [12, 16, 44, 14, 16, 16, 14, 26, 34]
        row_len = 9

    apply_common_sheet_layout(
        ws,
        title=title,
        subtitle="Historical rows are aggregated from real snapshot files in artifacts/sa-04-10/history.",
        header_row=4,
        table_name=path.stem.replace("-", "_"),
        rows=rows,
        widths=widths,
    )
    style_header(ws, 4, headers)

    for r_idx, row in enumerate(rows, 5):
        for c_idx, val in enumerate(row, 1):
            c = ws.cell(r_idx, c_idx, val)
            c.border = border
            c.alignment = Alignment(
                horizontal="left" if c_idx in {2, 3, row_len - 1, row_len} else "center",
                vertical="center",
                wrap_text=True,
            )
            if c_idx == 1 and val:
                c.number_format = "yyyy-mm-dd"

            # light color coding
            if kind != "security" and c_idx == 7:
                v = str(val).lower()
                if v in {"open", "high", "critical"}:
                    c.fill = PatternFill("solid", fgColor=RED)
                elif v == "clear":
                    c.fill = PatternFill("solid", fgColor=GREEN)
                elif v.startswith("pending") or v.startswith("historical"):
                    c.fill = PatternFill("solid", fgColor=AMBER)

            if kind == "security" and c_idx == 8:
                v = str(val).lower()
                if v in {"open", "fail", "error"}:
                    c.fill = PatternFill("solid", fgColor=RED)
                elif v in {"clear", "pass"}:
                    c.fill = PatternFill("solid", fgColor=GREEN)
                elif v.startswith("pending") or v.startswith("historical"):
                    c.fill = PatternFill("solid", fgColor=AMBER)

    summary = wb.create_sheet("Snapshot Summary")
    summary.sheet_view.showGridLines = False
    summary.merge_cells("A1:B1")
    summary["A1"] = f"{title} Summary"
    summary["A1"].font = Font(color=WHITE, bold=True, size=13)
    summary["A1"].fill = PatternFill("solid", fgColor=DARK)
    summary["A1"].alignment = Alignment(horizontal="center")

    summary_rows = snapshot_summary_rows(kind, snapshot)
    for i, (k, v) in enumerate(summary_rows, 3):
        summary.cell(i, 1, k).font = Font(bold=True)
        summary.cell(i, 1).fill = PatternFill("solid", fgColor=LIGHT)
        summary.cell(i, 1).border = border
        summary.cell(i, 2, v).border = border
        summary.cell(i, 2).alignment = Alignment(wrap_text=True)

    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 92

    notes = wb.create_sheet("Read Me")
    notes.sheet_view.showGridLines = False
    notes.merge_cells("A1:B1")
    notes["A1"] = f"{title} — Audit Notes"
    notes["A1"].font = Font(color=WHITE, bold=True, size=13)
    notes["A1"].fill = PatternFill("solid", fgColor=DARK)
    notes_rows = [
        ("Purpose", "30-day FedRAMP SA-04(10) log generated from actual historical snapshots."),
        ("Source", "artifacts/sa-04-10/history/history.jsonl and daily snapshot JSON files."),
        ("Behavior", "Only real history is rendered; missing days are not padded with placeholder rows."),
        ("Audit cue", "Use with the evidence binder and OSCAL SSP package."),
    ]
    for i, (k, v) in enumerate(notes_rows, 3):
        notes.cell(i, 1, k).font = Font(bold=True)
        notes.cell(i, 1).fill = PatternFill("solid", fgColor=LIGHT)
        notes.cell(i, 1).border = border
        notes.cell(i, 2, v).border = border
        notes.cell(i, 2).alignment = Alignment(wrap_text=True)

    notes.column_dimensions["A"].width = 18
    notes.column_dimensions["B"].width = 92

    wb.save(path)


def render_rows(history: List[Dict[str, Any]]) -> Dict[str, List[List[Any]]]:
    history = last_30_days(history)
    dep_rows: List[List[Any]] = []
    code_rows: List[List[Any]] = []
    sec_rows: List[List[Any]] = []

    for snap in history:
        day = snap.get("date") or (snap.get("generated_at") or "")[:10]
        scope = snap.get("scope", "")
        repo = snap.get("repository", "")
        results = snap.get("results", {})
        overall = snap.get("overall", {})

        dep = results.get("dependabot", {})
        code = results.get("code_scanning", {})
        sec = results.get("secret_scanning", {})

        dep_rows.append([
            day,
            scope,
            repo,
            dep.get("count", 0),
            dep.get("blocking_count", 0),
            "High" if dep.get("blocking_count", 0) else "Low",
            "Open" if dep.get("blocking_count", 0) else "Clear",
            snap.get("generated_at", ""),
            dep.get("skip_reason") or "Historical snapshot",
        ])

        code_rows.append([
            day,
            scope,
            repo,
            code.get("count", 0),
            code.get("blocking_count", 0),
            "High" if code.get("blocking_count", 0) else "Low",
            "Open" if code.get("blocking_count", 0) else "Clear",
            snap.get("generated_at", ""),
            code.get("skip_reason") or "Historical snapshot",
        ])

        sec_rows.append([
            day,
            scope,
            repo,
            code.get("count", 0),
            dep.get("count", 0),
            sec.get("count", 0),
            overall.get("blocking_count", 0),
            overall.get("status", "").title(),
            snap.get("generated_at", ""),
            "Historical snapshot",
        ])

    return {
        "dependabot": dep_rows,
        "codeql": code_rows,
        "security": sec_rows,
    }


def main() -> int:
    args = parse_args()
    input_dir = Path(args.input_dir)
    output_dir = Path(args.output_dir)

    output_dir.mkdir(parents=True, exist_ok=True)

    history = read_history(input_dir)
    if not history:
        raise SystemExit("No historical snapshots found in artifacts/sa-04-10")

    rows = render_rows(history)
    latest = history[-1]

    build_workbook(
        output_dir / "dependabot_30_day_log.xlsx",
        "Dependabot 30-Day Log",
        "dependabot",
        rows["dependabot"],
        latest,
    )

    build_workbook(
        output_dir / "security_30_day_log.xlsx",
        "Security 30-Day Log",
        "security",
        rows["security"],
        latest,
    )

    build_workbook(
        output_dir / "codeql_30_day_log.xlsx",
        "CodeQL 30-Day Log",
        "codeql",
        rows["codeql"],
        latest,
    )

    print(f"Spreadsheets generated in {output_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
