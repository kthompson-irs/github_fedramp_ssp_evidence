#!/usr/bin/env python3
"""Convert audit log findings into POA&M rows and export CSV, JSON, summary, and XLSX."""

from __future__ import annotations

import csv
import json
import os
from dataclasses import asdict, dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

GH_OWNER = os.getenv("GH_OWNER") or os.getenv("GITHUB_OWNER") or ""
GH_REPO = os.getenv("GH_REPO") or os.getenv("GITHUB_REPO") or ""
GH_ENTERPRISE_SLUG = os.getenv("GH_ENTERPRISE_SLUG") or os.getenv("GITHUB_ENTERPRISE") or ""

OUTPUT_DIR = Path(os.getenv("OUTPUT_DIR", "poam-output"))
AUDIT_FINDINGS_JSON = Path(os.getenv("AUDIT_FINDINGS_JSON", str(OUTPUT_DIR / "findings.json")))
AUDIT_SOURCE_ERRORS_JSON = Path(os.getenv("AUDIT_SOURCE_ERRORS_JSON", str(OUTPUT_DIR / "audit_source_errors.json")))
POAM_TEMPLATE_PATH = Path(os.getenv("POAM_TEMPLATE_PATH", "")).expanduser() if os.getenv("POAM_TEMPLATE_PATH") else None

OUTPUT_CSV = Path(os.getenv("OUTPUT_CSV", str(OUTPUT_DIR / "poam_github.csv")))
OUTPUT_JSON = Path(os.getenv("OUTPUT_JSON", str(OUTPUT_DIR / "poam_github.json")))
OUTPUT_SUMMARY = Path(os.getenv("OUTPUT_SUMMARY", str(OUTPUT_DIR / "poam_summary.json")))
OUTPUT_XLSX = Path(os.getenv("OUTPUT_XLSX", str(OUTPUT_DIR / "fedramp_poam_populated.xlsx")))

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


@dataclass
class PoamRow:
    poam_id: str
    weakness_name: str
    weakness_description: str
    source_identifying_weakness: str
    asset_identifier: str
    severity: str
    risk_rating: str
    date_identified: str
    scheduled_completion_date: str
    actual_completion_date: str
    status: str
    owner: str
    remediation_action: str
    source_url: str
    finding_category: str = ""
    finding_actor: str = ""
    finding_action: str = ""
    finding_timestamp: str = ""
    finding_reason: str = ""
    finding_raw_json: str = ""


def _today() -> str:
    return datetime.now(timezone.utc).date().isoformat()


def _plus_30_days() -> str:
    return (datetime.now(timezone.utc) + timedelta(days=30)).date().isoformat()


def _map_severity(sev: str) -> str:
    s = (sev or "").strip().lower()
    if s in {"critical", "high"}:
        return "High"
    if s in {"medium", "moderate"}:
        return "Moderate"
    return "Low"


def _coerce_timestamp(value: Any) -> str:
    if not value:
        return ""
    if isinstance(value, (int, float)):
        ts = float(value)
        if ts > 1_000_000_000_000:
            ts = ts / 1000.0
        return datetime.fromtimestamp(ts, tz=timezone.utc).isoformat()
    if isinstance(value, str):
        return value
    return str(value)


def _asset_identifier() -> str:
    if GH_OWNER and GH_REPO:
        return f"{GH_OWNER}/{GH_REPO}"
    if GH_OWNER:
        return GH_OWNER
    return "GitHub Audit Log"


def _load_findings(path: Path) -> Tuple[List[Dict[str, Any]], Optional[str]]:
    if not path.exists():
        return [], f"Finding file not found: {path}"

    try:
        if path.suffix.lower() == ".jsonl":
            findings: List[Dict[str, Any]] = []
            with path.open("r", encoding="utf-8") as f:
                for line_no, line in enumerate(f, 1):
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        findings.append(json.loads(line))
                    except json.JSONDecodeError as exc:
                        return [], f"Invalid JSON on line {line_no} of {path}: {exc}"
            return findings, None

        data = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(data, list):
            return data, None
        if isinstance(data, dict):
            if "findings" in data and isinstance(data["findings"], list):
                return data["findings"], None
            return [data], None

        return [], f"Unsupported findings file format: {path}"
    except Exception as exc:
        return [], f"Failed to load findings from {path}: {exc}"


def _load_source_errors(path: Path) -> List[str]:
    if not path.exists():
        return []
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(data, list):
            return [str(x) for x in data if str(x).strip()]
        if isinstance(data, dict):
            if "errors" in data and isinstance(data["errors"], list):
                return [str(x) for x in data["errors"] if str(x).strip()]
            return [json.dumps(data, ensure_ascii=False)]
        return [str(data)]
    except Exception as exc:
        return [f"Failed to read source errors from {path}: {exc}"]


def _remediation_for_category(category: str, action: str, reason: str) -> str:
    c = (category or "").strip().lower()
    a = (action or "").strip().lower()
    r = (reason or "").strip().lower()

    if c == "token_exposure" or "token" in r or "pat" in r:
        return "Investigate the exposed credential, rotate or revoke the token, and review repository and audit log activity."
    if c == "auth_failure" or ("failed" in r and any(x in r for x in ("token", "pat", "oauth", "ssh"))):
        return "Review repeated authentication failures, confirm whether they are expected, and investigate for misuse or brute force."
    if c == "privilege_or_security_change" or any(
        x in a for x in ("branch_protection", "visibility", "remove_required_reviews", "disable_secret_scanning", "disable_code_scanning")
    ):
        return "Review the security or privilege change, verify authorization, and restore required protections if needed."
    if c == "suspicious_content":
        return "Validate the event for malicious or policy-bypass activity and preserve evidence for investigation."
    if c == "internal_network_activity":
        return "Confirm the source context for the internal network activity and validate whether the event is expected."
    return "Review the audit log finding and implement the required remediation or compensating control."


def _weakness_name(category: str, action: str) -> str:
    cat = (category or "general_review").replace("_", " ").title()
    act = action or "audit event"
    return f"GitHub audit log finding: {cat} ({act})"[:120]


def _weakness_description(finding: Dict[str, Any]) -> str:
    parts = []
    if finding.get("timestamp"):
        parts.append(f"Timestamp: {finding.get('timestamp')}")
    if finding.get("actor"):
        parts.append(f"Actor: {finding.get('actor')}")
    if finding.get("action"):
        parts.append(f"Action: {finding.get('action')}")
    if finding.get("category"):
        parts.append(f"Category: {finding.get('category')}")
    if finding.get("reason"):
        parts.append(f"Reason: {finding.get('reason')}")
    return " | ".join(parts) if parts else "GitHub audit log finding."


def findings_to_poam_rows(findings: List[Dict[str, Any]]) -> List[PoamRow]:
    rows: List[PoamRow] = []
    today = _today()
    completion = _plus_30_days()
    asset = _asset_identifier()

    for idx, finding in enumerate(findings, 1):
        severity = _map_severity(str(finding.get("severity", "Low")))
        category = str(finding.get("category", "") or "general_review")
        actor = str(finding.get("actor", "") or "")
        action = str(finding.get("action", "") or "")
        timestamp = _coerce_timestamp(finding.get("timestamp", ""))
        reason = str(finding.get("reason", "") or "")
        raw_json = json.dumps(finding.get("raw", finding), ensure_ascii=False, default=str)

        rows.append(
            PoamRow(
                poam_id=f"AUD-{idx:05d}",
                weakness_name=_weakness_name(category, action),
                weakness_description=_weakness_description(finding),
                source_identifying_weakness="GitHub Audit Log / Automated Detection",
                asset_identifier=asset,
                severity=severity,
                risk_rating=severity,
                date_identified=today,
                scheduled_completion_date=completion,
                actual_completion_date="",
                status="Open",
                owner="Security Operations",
                remediation_action=_remediation_for_category(category, action, reason),
                source_url="",
                finding_category=category,
                finding_actor=actor,
                finding_action=action,
                finding_timestamp=timestamp,
                finding_reason=reason,
                finding_raw_json=raw_json,
            )
        )

    return rows


def _write_csv(path: Path, rows: List[PoamRow]) -> None:
    fields = list(asdict(rows[0]).keys()) if rows else list(
        asdict(PoamRow("", "", "", "", "", "", "", "", "", "", "", "", "", "")).keys()
    )
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow(asdict(row))


def _write_json(path: Path, rows: List[PoamRow]) -> None:
    path.write_text(json.dumps([asdict(r) for r in rows], indent=2, ensure_ascii=False), encoding="utf-8")


def _write_summary(path: Path, rows: List[PoamRow], source_errors: List[str], findings_file: Path) -> None:
    summary = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "owner": GH_OWNER,
        "repo": GH_REPO,
        "enterprise_slug": GH_ENTERPRISE_SLUG,
        "findings_file": str(findings_file),
        "total_count": len(rows),
        "high_count": sum(1 for r in rows if r.severity == "High"),
        "moderate_count": sum(1 for r in rows if r.severity == "Moderate"),
        "low_count": sum(1 for r in rows if r.severity == "Low"),
        "source_errors": source_errors,
        "high_rows": [asdict(r) for r in rows if r.severity == "High"],
    }
    path.write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")


def _style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    widths = {
        1: 12, 2: 30, 3: 50, 4: 30, 5: 24, 6: 10, 7: 10, 8: 14, 9: 18, 10: 18,
        11: 12, 12: 18, 13: 40, 14: 36, 15: 20, 16: 20, 17: 20, 18: 24, 19: 48, 20: 48
    }
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width


def _write_workbook(path: Path, rows: List[PoamRow], source_errors: List[str], findings_file: Path) -> None:
    if POAM_TEMPLATE_PATH and POAM_TEMPLATE_PATH.exists():
        wb = load_workbook(POAM_TEMPLATE_PATH)
    else:
        wb = Workbook()

    if "Summary" in wb.sheetnames:
        ws_summary = wb["Summary"]
        ws_summary.delete_rows(1, ws_summary.max_row)
    else:
        ws_summary = wb.create_sheet("Summary")

    if "POAM" in wb.sheetnames:
        ws_poam = wb["POAM"]
        ws_poam.delete_rows(1, ws_poam.max_row)
    else:
        ws_poam = wb.create_sheet("POAM")

    ws_summary["A1"] = "FedRAMP POA&M Export Summary"
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary["A3"] = "Generated At"
    ws_summary["B3"] = datetime.now(timezone.utc).isoformat()
    ws_summary["A4"] = "Owner"
    ws_summary["B4"] = GH_OWNER
    ws_summary["A5"] = "Repository"
    ws_summary["B5"] = GH_REPO
    ws_summary["A6"] = "Enterprise Slug"
    ws_summary["B6"] = GH_ENTERPRISE_SLUG
    ws_summary["A7"] = "Findings File"
    ws_summary["B7"] = str(findings_file)
    ws_summary["A9"] = "Total Findings"
    ws_summary["B9"] = len(rows)
    ws_summary["A10"] = "High Findings"
    ws_summary["B10"] = sum(1 for r in rows if r.severity == "High")
    ws_summary["A11"] = "Moderate Findings"
    ws_summary["B11"] = sum(1 for r in rows if r.severity == "Moderate")
    ws_summary["A12"] = "Low Findings"
    ws_summary["B12"] = sum(1 for r in rows if r.severity == "Low")
    ws_summary["A14"] = "Source Errors"

    if source_errors:
        for i, err in enumerate(source_errors, 15):
            ws_summary[f"A{i}"] = f"- {err}"
    else:
        ws_summary["A15"] = "(none)"

    ws_summary.column_dimensions["A"].width = 22
    ws_summary.column_dimensions["B"].width = 90

    headers = [
        "poam_id", "weakness_name", "weakness_description", "source_identifying_weakness", "asset_identifier",
        "severity", "risk_rating", "date_identified", "scheduled_completion_date", "actual_completion_date",
        "status", "owner", "remediation_action", "source_url", "finding_category", "finding_actor",
        "finding_action", "finding_timestamp", "finding_reason", "finding_raw_json"
    ]
    ws_poam.append(headers)
    for row in rows:
        d = asdict(row)
        ws_poam.append([d.get(h, "") for h in headers])

    _style_sheet(ws_poam)
    _style_sheet(ws_summary)
    wb.save(path)


def main() -> int:
    findings, err = _load_findings(AUDIT_FINDINGS_JSON)
    source_errors: List[str] = []
    if err:
        source_errors.append(err)

    source_errors.extend(_load_source_errors(AUDIT_SOURCE_ERRORS_JSON))

    rows = findings_to_poam_rows(findings) if findings else []

    _write_csv(OUTPUT_CSV, rows)
    _write_json(OUTPUT_JSON, rows)
    _write_summary(OUTPUT_SUMMARY, rows, source_errors, AUDIT_FINDINGS_JSON)
    _write_workbook(OUTPUT_XLSX, rows, source_errors, AUDIT_FINDINGS_JSON)

    print(f"Wrote {len(rows)} POA&M rows from {AUDIT_FINDINGS_JSON}")
    print(f"CSV: {OUTPUT_CSV}")
    print(f"JSON: {OUTPUT_JSON}")
    print(f"Workbook: {OUTPUT_XLSX}")
    if source_errors:
        print("Source errors:")
        for e in source_errors:
            print(f"- {e}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
