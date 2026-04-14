#!/usr/bin/env python3
"""
Build a FedRAMP ATO package from collected evidence.

Inputs:
  --evidence_dir   Folder produced by collector scripts and manual evidence drops
  --out_dir        Final package folder

Expected evidence layout (root-relative paths under --evidence_dir):
  docs/ATO_Letter.pdf
  docs/SSP.pdf
  docs/SAR.pdf
  docs/POAM.xlsx
  docs/GitHub_Risk_Assessment.pdf
  docs/GitHub_MFA_Screenshot.png
  docs/AWS_MFA_Screenshot.png
  docs/Azure_MFA_Screenshot.png
  docs/SIEM_Dashboard.png
  docs/CI_CD_Pipeline_Log.pdf
  docs/Vulnerability_Review.pdf

  org.json
  repos.json
  audit_log.json
  members.json
  outside_collaborators.json
  rulesets/org_rulesets.json
  security_managers.json (optional)

  repos/<repo>/repo.json
  repos/<repo>/rulesets.json
  repos/<repo>/secret_scanning_alerts.json
  repos/<repo>/code_scanning_alerts.json
  repos/<repo>/dependabot_alerts.json
  repos/<repo>/branch_protection.json
  repos/<repo>/collaborators.json
  repos/<repo>/teams.json

  aws/config.json
  aws/cloudtrail.json
  aws/iam_summary.json
  azure/policy.json
  azure/account.json
  k8s/roles.txt
  k8s/rolebindings.txt
  k8s/clusterroles.txt
  k8s/clusterrolebindings.txt
  k8s/networkpolicies.yaml
  terraform/plan.txt
  terraform/show.txt
  terraform/state_list.txt
  terraform/providers.txt

Outputs:
  compliance_package_index.pdf
  ato_executive_summary.pdf
  github_third_party_risk_assessment.pdf
  control_evidence_index.pdf
  poam_summary.pdf
  ca06_binder_index.md
  control_status.json
  pass_fail_score.json
  package_manifest.json
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import time
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None  # type: ignore

from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, Spacer, Table, TableStyle, SimpleDocTemplate


# ---------------------------
# Data model
# ---------------------------

@dataclass
class ControlCheck:
    control_id: str
    title: str
    evidence_paths: List[str]
    command: str
    poam_ids: List[str]
    notes: str = ""
    requires_nonempty_json: bool = False
    requires_all_repos: bool = False


@dataclass
class ControlResult:
    control_id: str
    title: str
    status: str
    score: int
    missing: List[str]
    evidence: List[str]
    command: str
    poam_ids: List[str]
    notes: str


# ---------------------------
# CLI
# ---------------------------

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build FedRAMP ATO package PDFs and binder.")
    parser.add_argument("--evidence_dir", default="ca06_evidence", help="Input evidence directory")
    parser.add_argument("--out_dir", default="ca06_package", help="Output package directory")
    return parser.parse_args()


# ---------------------------
# Helpers
# ---------------------------

def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def file_exists(root: Path, rel: str) -> bool:
    return (root / rel).exists()


def nonempty_text(root: Path, rel: str) -> bool:
    p = root / rel
    return p.exists() and p.is_file() and p.read_text(encoding="utf-8", errors="ignore").strip() != ""


def json_exists_and_nonempty(root: Path, rel: str) -> bool:
    p = root / rel
    if not p.exists() or not p.is_file():
        return False
    try:
        data = read_json(p)
    except Exception:
        return False
    if data is None:
        return False
    if isinstance(data, list):
        return len(data) > 0
    if isinstance(data, dict):
        return len(data) > 0
    if isinstance(data, str):
        return data.strip() != ""
    return True


def discover_repos(evidence_dir: Path) -> List[str]:
    repos_path = evidence_dir / "repos.json"
    if not repos_path.exists():
        return []
    try:
        payload = read_json(repos_path)
        if isinstance(payload, list):
            names = []
            for item in payload:
                if isinstance(item, dict) and "name" in item:
                    names.append(str(item["name"]))
            return names
    except Exception:
        return []
    return []


def collect_repo_paths(repos: Sequence[str], leaf: str) -> List[str]:
    return [f"repos/{repo}/{leaf}" for repo in repos]


def normalize_rows(rows: Sequence[Sequence[Any]]) -> List[List[str]]:
    normalized: List[List[str]] = []
    for row in rows:
        normalized.append([str(cell) for cell in row])
    return normalized


def p(text: str, style: str = "Body") -> Paragraph:
    styles = get_styles()
    return Paragraph(text.replace("\n", "<br/>"), styles[style])


# ---------------------------
# Styles and PDF helpers
# ---------------------------

def get_styles():
    styles = getSampleStyleSheet()
    if "Title2" not in styles:
        styles.add(
            ParagraphStyle(
                name="Title2",
                parent=styles["Title"],
                fontName="Helvetica-Bold",
                fontSize=20,
                leading=24,
                textColor=colors.HexColor("#17324d"),
                alignment=TA_LEFT,
                spaceAfter=10,
            )
        )
    if "Subtitle2" not in styles:
        styles.add(
            ParagraphStyle(
                name="Subtitle2",
                parent=styles["BodyText"],
                fontName="Helvetica",
                fontSize=10,
                leading=13,
                textColor=colors.HexColor("#4b5563"),
                alignment=TA_LEFT,
                spaceAfter=12,
            )
        )
    if "Section2" not in styles:
        styles.add(
            ParagraphStyle(
                name="Section2",
                parent=styles["Heading2"],
                fontName="Helvetica-Bold",
                fontSize=12,
                leading=15,
                textColor=colors.HexColor("#17324d"),
                spaceBefore=8,
                spaceAfter=6,
            )
        )
    if "Body2" not in styles:
        styles.add(
            ParagraphStyle(
                name="Body2",
                parent=styles["BodyText"],
                fontName="Helvetica",
                fontSize=9.5,
                leading=13,
                spaceAfter=6,
            )
        )
    if "Small2" not in styles:
        styles.add(
            ParagraphStyle(
                name="Small2",
                parent=styles["BodyText"],
                fontName="Helvetica",
                fontSize=8.3,
                leading=10.5,
                textColor=colors.HexColor("#374151"),
            )
        )
    if "Note2" not in styles:
        styles.add(
            ParagraphStyle(
                name="Note2",
                parent=styles["BodyText"],
                fontName="Helvetica-Oblique",
                fontSize=8.5,
                leading=11,
                textColor=colors.HexColor("#6b7280"),
            )
        )
    return styles


def header_footer(canvas, doc):
    canvas.saveState()
    w, h = LETTER
    canvas.setStrokeColor(colors.HexColor("#d1d5db"))
    canvas.setLineWidth(0.5)
    canvas.line(doc.leftMargin, h - 0.55 * inch, w - doc.rightMargin, h - 0.55 * inch)
    canvas.line(doc.leftMargin, 0.62 * inch, w - doc.rightMargin, 0.62 * inch)
    canvas.setFont("Helvetica-Bold", 8.5)
    canvas.setFillColor(colors.HexColor("#17324d"))
    canvas.drawString(doc.leftMargin, h - 0.42 * inch, "FedRAMP CA-6 ATO Compliance Package")
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(colors.HexColor("#6b7280"))
    canvas.drawRightString(w - doc.rightMargin, h - 0.42 * inch, f"Page {doc.page}")
    canvas.drawString(
        doc.leftMargin,
        0.38 * inch,
        "Generated by ca06_build_fedramp_ato_package.py",
    )
    canvas.restoreState()


def make_title_block(title: str, subtitle: str):
    return [
        Spacer(1, 0.08 * inch),
        Paragraph(title, get_styles()["Title2"]),
        Paragraph(subtitle, get_styles()["Subtitle2"]),
        Spacer(1, 0.07 * inch),
    ]


def table(rows: Sequence[Sequence[Any]], col_widths: Sequence[float]) -> Table:
    rows = normalize_rows(rows)
    t = Table(rows, colWidths=col_widths, hAlign="LEFT", repeatRows=1)
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#17324d")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8.2),
                ("LEADING", (0, 0), (-1, -1), 10.2),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#9ca3af")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor("#f8fafc")]),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    return t


def build_pdf(path: Path, title: str, subtitle: str, sections: List[Tuple[str, str]], tables: List[Tuple[List[List[Any]], List[float]]] = None):
    tables = tables or []
    doc = SimpleDocTemplate(
        str(path),
        pagesize=LETTER,
        leftMargin=0.75 * inch,
        rightMargin=0.75 * inch,
        topMargin=0.9 * inch,
        bottomMargin=0.75 * inch,
    )
    story: List[Any] = []
    story += make_title_block(title, subtitle)
    for heading, body in sections:
        story.append(Paragraph(heading, get_styles()["Section2"]))
        story.append(Paragraph(body, get_styles()["Body2"]))
    for rows, widths in tables:
        story.append(Spacer(1, 0.08 * inch))
        story.append(table(rows, widths))
    doc.build(story, onFirstPage=header_footer, onLaterPages=header_footer)


# ---------------------------
# POA&M parsing
# ---------------------------

def parse_poam_xlsx(path: Path) -> List[Dict[str, str]]:
    if load_workbook is None or not path.exists():
        return []
    try:
        wb = load_workbook(path, data_only=True)
    except Exception:
        return []
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    idx = {name: i for i, name in enumerate(headers)}

    def get(row, *names):
        for n in names:
            i = idx.get(n)
            if i is not None and i < len(row):
                v = row[i]
                if v is not None and str(v).strip() != "":
                    return str(v).strip()
        return ""

    parsed: List[Dict[str, str]] = []
    for row in rows[1:]:
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue
        parsed.append(
            {
                "ID": get(row, "id", "poam id", "poam_id"),
                "Control": get(row, "control", "control id", "control_id"),
                "Weakness": get(row, "weakness", "finding", "issue"),
                "Severity": get(row, "severity"),
                "Owner": get(row, "owner", "responsible party"),
                "Due Date": get(row, "due date", "due", "remediation due date"),
                "Status": get(row, "status"),
            }
        )
    return parsed


# ---------------------------
# Control checks
# ---------------------------

def build_checks(evidence_dir: Path, repos: List[str]) -> List[ControlCheck]:
    repo_json_paths = collect_repo_paths(repos, "repo.json")
    repo_ruleset_paths = collect_repo_paths(repos, "rulesets.json")
    repo_secret_paths = collect_repo_paths(repos, "secret_scanning_alerts.json")
    repo_code_paths = collect_repo_paths(repos, "code_scanning_alerts.json")
    repo_dep_paths = collect_repo_paths(repos, "dependabot_alerts.json")
    repo_branch_paths = collect_repo_paths(repos, "branch_protection.json")
    repo_collab_paths = collect_repo_paths(repos, "collaborators.json")
    repo_team_paths = collect_repo_paths(repos, "teams.json")

    checks = [
        ControlCheck(
            control_id="CA-6",
            title="Authorization",
            evidence_paths=[
                "docs/ATO_Letter.pdf",
                "docs/SSP.pdf",
                "docs/SAR.pdf",
                "docs/POAM.xlsx",
                "docs/GitHub_Risk_Assessment.pdf",
            ],
            command="Document review of ATO letter, SSP, SAR, POA&M, and GitHub risk assessment.",
            poam_ids=[],
            notes="Primary authorization evidence.",
        ),
        ControlCheck(
            control_id="SA-9",
            title="External Services (GitHub)",
            evidence_paths=[
                "org.json",
                "repos.json",
                "audit_log.json",
                "members.json",
                "outside_collaborators.json",
                "rulesets/org_rulesets.json",
            ]
            + repo_json_paths
            + repo_ruleset_paths
            + repo_secret_paths
            + repo_code_paths
            + repo_dep_paths
            + repo_branch_paths
            + repo_collab_paths
            + repo_team_paths,
            command="gh api /orgs/<ORG> ; gh api /orgs/<ORG>/repos ; gh api /orgs/<ORG>/audit-log",
            poam_ids=["GH-01"],
            notes="GitHub treated as external service; repo-level evidence required.",
            requires_nonempty_json=True,
            requires_all_repos=True,
        ),
        ControlCheck(
            control_id="AC-2",
            title="Account Management",
            evidence_paths=["members.json", "outside_collaborators.json"],
            command="gh api /orgs/<ORG>/members ; gh api /orgs/<ORG>/outside_collaborators",
            poam_ids=["GH-02"],
            notes="Org membership and collaborator inventory.",
            requires_nonempty_json=True,
        ),
        ControlCheck(
            control_id="AC-3",
            title="Access Enforcement",
            evidence_paths=repo_ruleset_paths + repo_branch_paths,
            command="gh api /repos/<ORG>/<REPO>/rulesets ; gh api /repos/<ORG>/<REPO>/branches/<default>/protection",
            poam_ids=["GH-03"],
            notes="Branch rulesets and protections across all repositories.",
            requires_all_repos=True,
        ),
        ControlCheck(
            control_id="AC-6",
            title="Least Privilege",
            evidence_paths=repo_collab_paths + repo_team_paths,
            command="gh api /repos/<ORG>/<REPO>/collaborators ; gh api /repos/<ORG>/<REPO>/teams",
            poam_ids=["GH-03"],
            notes="Direct collaborators and team access must be limited.",
            requires_all_repos=True,
        ),
        ControlCheck(
            control_id="IA-2",
            title="Identification and Authentication",
            evidence_paths=[
                "docs/GitHub_MFA_Screenshot.png",
                "docs/AWS_MFA_Screenshot.png",
                "docs/Azure_MFA_Screenshot.png",
                "org.json",
                "aws/iam_summary.json",
                "azure/account.json",
            ],
            command="Manual screenshot review plus cloud identity summaries.",
            poam_ids=["GH-04"],
            notes="MFA evidence is required from GitHub and cloud identity systems.",
        ),
        ControlCheck(
            control_id="AU-2/AU-6",
            title="Audit and Accountability",
            evidence_paths=["audit_log.json", "docs/SIEM_Dashboard.png", "docs/Audit_Review_Notes.pdf"],
            command="gh api /orgs/<ORG>/audit-log",
            poam_ids=["GH-05"],
            notes="Audit logs must exist and be reviewed in SIEM.",
            requires_nonempty_json=True,
        ),
        ControlCheck(
            control_id="CM-2/CM-6",
            title="Configuration Management",
            evidence_paths=[
                "terraform/plan.txt",
                "terraform/show.txt",
                "terraform/state_list.txt",
                "terraform/providers.txt",
                "aws/config.json",
                "azure/policy.json",
            ],
            command="terraform plan ; terraform show ; terraform state list ; AWS Config / Azure Policy exports",
            poam_ids=["GH-06"],
            notes="Infrastructure baseline and policy compliance.",
        ),
        ControlCheck(
            control_id="RA-5",
            title="Vulnerability Scanning",
            evidence_paths=[
                "docs/Vulnerability_Review.pdf",
                "repos.json",
                "rulesets/org_rulesets.json",
            ]
            + repo_secret_paths
            + repo_code_paths
            + repo_dep_paths,
            command="gh api /repos/<ORG>/<REPO>/secret-scanning/alerts ; gh api /repos/<ORG>/<REPO>/code-scanning/alerts ; gh api /repos/<ORG>/<REPO>/dependabot/alerts",
            poam_ids=["GH-07"],
            notes="Dependency and code scanning outputs must be present.",
            requires_all_repos=True,
        ),
        ControlCheck(
            control_id="SC-7",
            title="Boundary Protection",
            evidence_paths=["k8s/networkpolicies.yaml", "aws/cloudtrail.json", "azure/policy.json"],
            command="kubectl get networkpolicies -A -o yaml ; cloud boundary policy exports",
            poam_ids=["GH-08"],
            notes="Network restrictions and cloud boundary evidence.",
        ),
        ControlCheck(
            control_id="SI-2",
            title="Flaw Remediation / Patch Management",
            evidence_paths=["docs/CI_CD_Pipeline_Log.pdf", "docs/Vulnerability_Review.pdf"],
            command="CI/CD pipeline logs showing patch and rebuild flow",
            poam_ids=["GH-09"],
            notes="Patch and remediation cadence must be documented.",
        ),
        ControlCheck(
            control_id="K8S",
            title="Kubernetes Security Posture",
            evidence_paths=[
                "k8s/roles.txt",
                "k8s/rolebindings.txt",
                "k8s/clusterroles.txt",
                "k8s/clusterrolebindings.txt",
                "k8s/networkpolicies.yaml",
            ],
            command="kubectl get roles -A ; kubectl get rolebindings -A ; kubectl get clusterroles ; kubectl get clusterrolebindings ; kubectl get networkpolicies -A -o yaml",
            poam_ids=["GH-10"],
            notes="RBAC and network policies must be explicit.",
        ),
        ControlCheck(
            control_id="LOG",
            title="Logging Centralization",
            evidence_paths=["docs/SIEM_Dashboard.png", "audit_log.json", "aws/cloudtrail.json", "azure/policy.json"],
            command="SIEM export plus GitHub / cloud log sources",
            poam_ids=["GH-05"],
            notes="GitHub, cloud, and security logs must be centralized.",
            requires_nonempty_json=True,
        ),
    ]
    return checks


def evaluate_check(evidence_dir: Path, check: ControlCheck, repos: List[str]) -> ControlResult:
    missing: List[str] = []

    def file_ok(rel: str) -> bool:
        return file_exists(evidence_dir, rel)

    # Required evidence paths
    for rel in check.evidence_paths:
        if rel.endswith("/"):
            continue
        if not file_ok(rel):
            missing.append(rel)

    # Special rule: if the control requires all repos, make sure every repo has its core files.
    if check.requires_all_repos:
        required_repo_sets = {
            "SA-9": [
                "repo.json",
                "rulesets.json",
                "secret_scanning_alerts.json",
                "code_scanning_alerts.json",
                "dependabot_alerts.json",
                "branch_protection.json",
                "collaborators.json",
                "teams.json",
            ],
            "AC-3": ["rulesets.json", "branch_protection.json"],
            "AC-6": ["collaborators.json", "teams.json"],
            "RA-5": ["secret_scanning_alerts.json", "code_scanning_alerts.json", "dependabot_alerts.json"],
        }
        leaves = required_repo_sets.get(check.control_id, [])
        for repo in repos:
            for leaf in leaves:
                rel = f"repos/{repo}/{leaf}"
                if not file_ok(rel):
                    missing.append(rel)

    # JSON must not be empty when flagged
    if check.requires_nonempty_json:
        json_paths = [p for p in check.evidence_paths if p.endswith(".json")]
        for rel in json_paths:
            if not json_exists_and_nonempty(evidence_dir, rel):
                if rel not in missing:
                    missing.append(rel)

    status = "PASS" if not missing else "FAIL"
    score = 100 if status == "PASS" else 0
    return ControlResult(
        control_id=check.control_id,
        title=check.title,
        status=status,
        score=score,
        missing=sorted(set(missing)),
        evidence=check.evidence_paths,
        command=check.command,
        poam_ids=check.poam_ids,
        notes=check.notes,
    )


# ---------------------------
# PDF report generation
# ---------------------------

def generate_compliance_package_index(out_dir: Path, results: List[ControlResult], score: float):
    build_pdf(
        out_dir / "compliance_package_index.pdf",
        "FedRAMP CA-6 ATO Compliance Package",
        "Package index and readiness summary for GitHub + cloud evidence.",
        [
            ("Scope", "This package is built from collected GitHub evidence, cloud evidence, Kubernetes evidence, Terraform evidence, and manual review artifacts."),
            ("Readiness Score", f"Overall pass rate: <b>{score:.1f}%</b>. The package is PASS only when all required controls pass and all required artifacts exist."),
        ],
        [
            (
                [
                    ["Control", "Status", "POA&M", "Evidence Shortlist"],
                    *[
                        [
                            r.control_id,
                            r.status,
                            ", ".join(r.poam_ids) if r.poam_ids else "N/A",
                            ", ".join(r.evidence[:3]) + (" ..." if len(r.evidence) > 3 else ""),
                        ]
                        for r in results
                    ],
                ],
                [0.7 * inch, 0.65 * inch, 0.9 * inch, 4.9 * inch],
            )
        ],
    )


def generate_executive_summary(out_dir: Path, results: List[ControlResult], score: float):
    build_pdf(
        out_dir / "ato_executive_summary.pdf",
        "ATO Executive Summary",
        "Authorization summary for a multi-cloud production environment using GitHub.com as an external service.",
        [
            ("System Scope", "The system uses GitHub.com for source control and CI/CD support, AWS GovCloud (US) and Azure Government for production hosting, Terraform for infrastructure provisioning, and Kubernetes for workload orchestration."),
            ("Authorization Basis", "Authorization is supported by the ATO letter, SSP, SAR, and POA&M. GitHub.com is treated as an external service under SA-9 and must be risk-accepted explicitly in the authorization package."),
            ("Critical Notes", "No secrets, production data, or controlled information should be stored in GitHub. Branch protections, rulesets, MFA, logging, and scanning must be present and evidenced."),
            ("Current Pass Rate", f"Current package pass rate: <b>{score:.1f}%</b>. See the binder for control-by-control results."),
        ],
    )


def generate_github_risk_assessment(out_dir: Path, results: List[ControlResult]):
    rows = [["Risk ID", "Risk", "Impact", "Mitigation"]]
    rows += [
        ["GH-01", "Sensitive data exposure in repositories", "High", "Private repos, no CUI/secrets, data classification"],
        ["GH-02", "Weak access control or stale collaborators", "High", "MFA, SSO, periodic membership review"],
        ["GH-03", "Secret leakage in commits or workflows", "High", "Secret scanning, push protection, external secret store"],
        ["GH-04", "Supply chain compromise through code changes", "High", "Branch protection, required reviews, signed changes, scanning"],
        ["GH-05", "Insufficient audit visibility", "Medium", "Audit-log export to SIEM and recurring review"],
    ]
    build_pdf(
        out_dir / "github_third_party_risk_assessment.pdf",
        "GitHub.com Third-Party Risk Assessment",
        "External service assessment for SA-9 support in the FedRAMP authorization package.",
        [
            ("Service Overview", "GitHub.com is used for source code management and CI/CD support. It is treated as an external service and is not inside the federal cloud authorization boundary."),
            ("Residual Risk Statement", "Residual risk is acceptable only when GitHub remains restricted to non-sensitive source code and IaC, security controls are enabled and monitored, and the AO explicitly accepts the external-service risk in the authorization package."),
        ],
        [(rows, [0.7 * inch, 2.9 * inch, 0.9 * inch, 2.3 * inch])],
    )


def generate_binder_pdf(out_dir: Path, results: List[ControlResult]):
    rows = [["Control", "Status", "Evidence Files", "Command", "POA&M"]]
    for r in results:
        rows.append(
            [
                r.control_id,
                r.status,
                "<br/>".join(r.evidence[:6]) + ("<br/>..." if len(r.evidence) > 6 else ""),
                r.command,
                ", ".join(r.poam_ids) if r.poam_ids else "N/A",
            ]
        )
    build_pdf(
        out_dir / "control_evidence_index.pdf",
        "Control Evidence Index",
        "Control-to-evidence mapping with exact filenames, commands, and POA&M references.",
        [
            ("How to Use", "For each control, open the exact file listed, run the command if needed, and point to the matching POA&M ID for any gap."),
        ],
        [(rows, [0.65 * inch, 0.55 * inch, 2.45 * inch, 2.65 * inch, 0.7 * inch])],
    )


def generate_poam_pdf(out_dir: Path, poam_rows: List[Dict[str, str]], results: List[ControlResult]):
    if not poam_rows:
        poam_rows = []
        for r in results:
            if r.status == "FAIL":
                poam_rows.append(
                    {
                        "ID": ", ".join(r.poam_ids) if r.poam_ids else "",
                        "Control": r.control_id,
                        "Weakness": f"Missing or incomplete evidence: {', '.join(r.missing[:3])}" + (" ..." if len(r.missing) > 3 else ""),
                        "Severity": "Medium" if r.control_id not in {"CA-6", "IA-2"} else "High",
                        "Owner": "Security / Platform",
                        "Due Date": "TBD",
                        "Status": "Open",
                    }
                )

    rows = [["ID", "Control", "Weakness", "Severity", "Owner", "Due Date", "Status"]]
    for item in poam_rows:
        rows.append(
            [
                item.get("ID", ""),
                item.get("Control", ""),
                item.get("Weakness", ""),
                item.get("Severity", ""),
                item.get("Owner", ""),
                item.get("Due Date", ""),
                item.get("Status", ""),
            ]
        )

    build_pdf(
        out_dir / "poam_summary.pdf",
        "POA&M Summary",
        "Remediation tracker for open findings and missing evidence.",
        [
            ("Note", "If a POA&M spreadsheet exists in the evidence folder, it is parsed and summarized here. Otherwise, open items are derived from failed controls in this run."),
        ],
        [(rows, [0.55 * inch, 0.7 * inch, 2.95 * inch, 0.65 * inch, 0.85 * inch, 0.65 * inch, 0.55 * inch])],
    )


# ---------------------------
# Binder markdown / scoring
# ---------------------------

def make_binder_markdown(out_dir: Path, results: List[ControlResult], score: float, repos: List[str]) -> str:
    lines = [
        "# FedRAMP ATO Binder Index",
        "",
        f"Generated UTC: {time.strftime('%Y-%m-%dT%H:%M:%SZ', time.gmtime())}",
        f"Overall pass rate: {score:.1f}%",
        "",
        "| Control | Status | POA&M IDs | Evidence files (short) | Command |",
        "|---|---:|---|---|---|",
    ]
    for r in results:
        evidence_short = "<br>".join(r.evidence[:4]) + ("<br>..." if len(r.evidence) > 4 else "")
        lines.append(
            f"| {r.control_id} | {r.status} | {', '.join(r.poam_ids) if r.poam_ids else 'N/A'} | {evidence_short} | {r.command} |"
        )
    lines += [
        "",
        "## Repository Scope",
        "",
        ", ".join(repos) if repos else "No repositories discovered.",
        "",
        "## Notes",
        "",
        "- File names in the binder should match evidence files exactly.",
        "- Any FAIL row should have a linked POA&M entry.",
    ]
    return "\n".join(lines) + "\n"


# ---------------------------
# Main
# ---------------------------

def main() -> int:
    args = parse_args()
    evidence_dir = Path(args.evidence_dir).resolve()
    out_dir = Path(args.out_dir).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    if not evidence_dir.exists():
        print(f"ERROR: evidence_dir does not exist: {evidence_dir}", file=sys.stderr)
        return 2

    repos = discover_repos(evidence_dir)
    checks = build_checks(evidence_dir, repos)
    results = [evaluate_check(evidence_dir, chk, repos) for chk in checks]

    pass_count = sum(1 for r in results if r.status == "PASS")
    total = len(results)
    score = (pass_count / total * 100.0) if total else 0.0
    overall = "PASS" if pass_count == total and total > 0 else "FAIL"

    poam_rows = parse_poam_xlsx(evidence_dir / "docs" / "POAM.xlsx")

    # Write JSON outputs.
    write_json(out_dir / "control_status.json", [asdict(r) for r in results])
    write_json(
        out_dir / "pass_fail_score.json",
        {
            "generated_utc": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
            "overall": overall,
            "score_percent": round(score, 1),
            "passed_controls": pass_count,
            "total_controls": total,
            "repos_discovered": repos,
        },
    )
    write_json(
        out_dir / "package_manifest.json",
        {
            "generated_utc": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
            "inputs": {
                "evidence_dir": str(evidence_dir),
            },
            "outputs": [
                "compliance_package_index.pdf",
                "ato_executive_summary.pdf",
                "github_third_party_risk_assessment.pdf",
                "control_evidence_index.pdf",
                "poam_summary.pdf",
                "ca06_binder_index.md",
                "control_status.json",
                "pass_fail_score.json",
            ],
            "overall": overall,
            "score_percent": round(score, 1),
        },
    )

    # Markdown binder.
    binder_md = make_binder_markdown(out_dir, results, score, repos)
    write_text(out_dir / "ca06_binder_index.md", binder_md)

    # PDFs.
    generate_compliance_package_index(out_dir, results, score)
    generate_executive_summary(out_dir, results, score)
    generate_github_risk_assessment(out_dir, results)
    generate_binder_pdf(out_dir, results)
    generate_poam_pdf(out_dir, poam_rows, results)

    print(f"Package written to: {out_dir}")
    print(f"Overall: {overall} ({score:.1f}%)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
